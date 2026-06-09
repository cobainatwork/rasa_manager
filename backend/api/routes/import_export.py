"""
Excel 批次匯入 / 匯出路由。

匯入規格（§4.3 + §21.14）：
  - 僅接受 .xlsx，上限 10 MB / 5000 行
  - 必填欄：question、answer、category_path（/ 分隔）
  - 選填欄：tags（逗號分隔字串）
  - Superadmin 匯入時直接建立為 approved（與單筆建立行為一致）；其他使用者建立為 draft
  - category_path 不存在時自動建立節點
  - mode=append（預設）：相同 agent_id + question 已存在 → 跳過
  - mode=replace：先刪除該 agent 所有 FAQ 再匯入（全量取代）
  - 操作結束後關閉 workbook（openpyxl read_only 模式）

匯出規格（§4.3）：
  - 欄位：id、status、category_path、tags、question、answer、version、created_at、updated_at
  - 回傳 StreamingResponse（xlsx）
"""
from __future__ import annotations

import io
import re
import uuid
from datetime import datetime
from typing import Any, Literal, Optional
from urllib.parse import quote

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from api.database.models import Agent, Category, KnowledgeItem, User
from api.database.session import get_db
from api.dependencies import _get_redis, get_accessible_agent, get_current_user
from api.services.audit import record_audit
from api.utils.category_path import build_category_path, collect_category_subtree

router = APIRouter(tags=["import-export"])

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
MAX_ROWS = 5000


def _sanitize_for_filename(scope: str) -> str:
    """將分類路徑轉為可用於檔名的字串（去除或替換不合法字元）。"""
    # 把 / 換成 _（分類路徑分隔符號）
    sanitized = scope.replace("/", "_")
    # 移除 Windows/Linux 都禁用的特殊字元
    sanitized = re.sub(r'[\\:*?"<>|]', "", sanitized)
    # 連續底線合併、去頭尾底線
    sanitized = re.sub(r"_+", "_", sanitized).strip("_")
    # 長度上限 50 字元
    return sanitized[:50] if sanitized else "未命名"


def _generate_export_filename(scope: str, agent_id: str) -> str:
    """產生含每日遞增序號的匯出檔名，例如 全量_export_260504_00.xlsx。"""
    sanitized = _sanitize_for_filename(scope)
    yymmdd = datetime.now().strftime("%y%m%d")
    try:
        r = _get_redis()
        key = f"export:counter:{agent_id}:{scope}:{yymmdd}"
        count = int(r.incr(key)) - 1          # 第一次 incr=1 → count=0 → "00"
        r.expire(key, 86400 * 7)               # 7 天後過期（日期在 key 中，不會誤重置）
        nn = f"{count:02d}"
    except Exception:
        # Redis 不可用時以秒級時間戳作 fallback（極端並發下仍可能重複，但可接受）
        nn = datetime.now().strftime("%H%M%S")
    return f"{sanitized}_export_{yymmdd}_{nn}.xlsx"


def _get_or_create_default_category(db: Session, agent_id: Any) -> Any:
    """取得或建立「未分類」根分類，回傳 category.id。"""
    DEFAULT_NAME = "未分類"
    cat = (
        db.query(Category)
        .filter(
            Category.agent_id == agent_id,
            Category.name == DEFAULT_NAME,
            Category.parent_id.is_(None),
        )
        .first()
    )
    if cat is None:
        cat = Category(
            id=uuid.uuid4(),
            agent_id=agent_id,
            parent_id=None,
            name=DEFAULT_NAME,
            sort_order=0,
        )
        db.add(cat)
        db.flush()
    return cat.id


def _preload_category_index(
    db: Session, agent_id: Any
) -> dict[tuple[Optional[Any], str], Any]:
    """預載指定 agent 的全部分類，回傳 (parent_id, name) → category.id 的對照表。

    用於匯入大量 FAQ 時避免每行每層 category_path 都打一次 SELECT；
    可達成 O(rows × depth) → O(1) 預載 + dict lookup 的轉換。

    對非預期 row 形態具備容錯（測試 mock 可能回非 3-tuple）：解構失敗的列略過，
    僅放棄該列在 cache 內的命中（後續會回退到 DB 查詢路徑），不影響整體正確性。
    """
    rows = (
        db.query(Category.id, Category.parent_id, Category.name)
        .filter(Category.agent_id == agent_id)
        .all()
    )
    index: dict[tuple[Optional[Any], str], Any] = {}
    for row in rows:
        try:
            cid, parent_id, name = row
        except (TypeError, ValueError):
            continue
        index[(parent_id, str(name))] = cid
    return index


def _resolve_category_path(
    db: Session,
    agent_id: Any,
    path_str: str,
    cache: Optional[dict[tuple[Optional[Any], str], Any]] = None,
) -> tuple[Any, bool]:
    """
    解析 / 分隔的 category_path，自動建立缺少的節點。
    回傳 (最末層 category.id, 本次是否新建任何分類層)。

    若提供 cache（由 _preload_category_index 建立），優先以 dict lookup 解析；
    cache miss 才回退到 DB 查詢，並把建立的新節點回填 cache 供後續 row 共用。
    維持原本 SAVEPOINT 內 db.flush() 的 semantic（建立分類後仍 flush 取得 id）。
    """
    parts = [p.strip() for p in path_str.split("/") if p.strip()]
    if not parts:
        raise ValueError(f"category_path 不可為空：{path_str!r}")

    parent_id: Optional[Any] = None
    created = False
    for part in parts:
        cat_id: Optional[Any] = None
        key = (parent_id, part)
        if cache is not None and key in cache:
            cat_id = cache[key]
        else:
            cat = (
                db.query(Category)
                .filter(
                    Category.agent_id == agent_id,
                    Category.name == part,
                    Category.parent_id == parent_id,
                )
                .first()
            )
            if cat is not None:
                cat_id = cat.id
                if cache is not None:
                    cache[key] = cat_id
        if cat_id is None:
            new_cat = Category(
                id=uuid.uuid4(),
                agent_id=agent_id,
                parent_id=parent_id,
                name=part,
                sort_order=0,
            )
            db.add(new_cat)
            db.flush()
            cat_id = new_cat.id
            if cache is not None:
                cache[key] = cat_id
            created = True
        parent_id = cat_id

    return parent_id, created


# ── 匯入端點 ──────────────────────────────────────────────────────────────────

@router.post("/api/v1/agents/{agent_id}/faqs/import")
def import_faqs(
    agent_id: uuid.UUID,
    file: UploadFile = File(...),
    mode: Literal["append", "replace"] = Query("append", description="append=跳過重複；replace=先清空再匯入"),
    access: tuple[Agent, str | None] = Depends(get_accessible_agent),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    del access  # 僅做存取驗證

    # ── 檔案格式與大小驗證 ─────────────────────────────────────────────────────
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="僅接受 .xlsx 格式檔案")

    content = file.file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="檔案大小超過 10 MB 上限")

    # ── 解析 Excel ─────────────────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("找不到工作表")
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{exc}") from exc

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 無資料列（需含標題行）")

    header = [str(c).strip().lower() if c is not None else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"資料列數超過 {MAX_ROWS} 行上限")

    # ── 欄位索引解析 ───────────────────────────────────────────────────────────
    def require_col(name: str) -> int:
        if name not in header:
            raise HTTPException(status_code=400, detail=f"缺少必填欄位：{name}")
        return header.index(name)

    q_idx = require_col("question")
    a_idx = require_col("answer")
    # category_path 為選填：有則解析路徑，無則歸入「未分類」
    cp_idx: Optional[int] = header.index("category_path") if "category_path" in header else None
    tags_idx: Optional[int] = header.index("tags") if "tags" in header else None

    # ── replace 模式：先刪除該 agent 所有 FAQ ────────────────────────────────
    if mode == "replace":
        db.query(KnowledgeItem).filter(KnowledgeItem.agent_id == agent_id).delete(
            synchronize_session="fetch"
        )
        db.flush()

    # ── 預載已存在的 question（避免重複；replace 模式剛清空，故為空集合）───────
    existing_questions: set[str] = (
        set()
        if mode == "replace"
        else {
            str(q)
            for (q,) in db.query(KnowledgeItem.question)
            .filter(KnowledgeItem.agent_id == agent_id)
            .all()
        }
    )

    # ── 預載分類索引：避免每行每層 category_path 都打一次 SELECT ────────────────
    # 5000 行 × 2 層深 → 從 10000 次 SELECT 縮成 1 次預載 + dict lookup。
    category_cache: dict[tuple[Optional[Any], str], Any] = _preload_category_index(
        db, agent_id
    )

    # Superadmin 匯入直接核准（與 create_faq 行為一致），其他使用者建立為 draft
    initial_status = "approved" if current_user.is_superadmin else "draft"

    success = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    new_categories: set[str] = set()

    # 「未分類」根分類 ID 的迴圈層快取（避免在 SAVEPOINT 內建立分類）
    _default_cat_id: Optional[Any] = None

    for row_num, row in enumerate(data_rows, start=2):
        def cell(idx: int) -> str:
            val = row[idx] if idx < len(row) else None  # type: ignore[index]
            return str(val).strip() if val is not None else ""

        question = cell(q_idx)
        answer = cell(a_idx)
        category_path_val = cell(cp_idx) if cp_idx is not None else ""
        tags_raw = cell(tags_idx) if tags_idx is not None else ""

        # 空列跳過（question 與 answer 均空才算空列）
        if not question and not answer:
            continue

        if not question or not answer:
            errors.append({
                "row": row_num,
                "reason": "question / answer 不可為空",
            })
            continue

        if question in existing_questions:
            skipped += 1
            continue

        tags_list = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

        # 使用 SAVEPOINT（nested transaction）：單筆失敗只 rollback 該筆，
        # 不影響先前已 flush 的成功項目。
        # category_path 選填：有則留待 begin_nested 內解析；無則在 begin_nested 外部取得預設分類
        if not category_path_val:
            if _default_cat_id is None:
                _default_cat_id = _get_or_create_default_category(db, agent_id)
                db.flush()  # 確保 Category 進入 DB（在任何 begin_nested 之前）

        try:
            with db.begin_nested():
                # category_path 選填：有則解析路徑（自動建立缺失節點），無則歸入「未分類」
                if category_path_val:
                    category_id, created = _resolve_category_path(
                        db, agent_id, category_path_val, cache=category_cache
                    )
                    if created:
                        new_categories.add(category_path_val)
                else:
                    category_id = _default_cat_id
                    created = False

                item = KnowledgeItem(
                    id=uuid.uuid4(),
                    agent_id=agent_id,
                    category_id=category_id,
                    question=question,
                    answer=answer,
                    tags=tags_list,
                    status=initial_status,
                    version=1,
                    created_by=current_user.id,
                )
                db.add(item)
                db.flush()

                record_audit(
                    db,
                    agent_id=agent_id,
                    item_id=item.id,
                    action="import",
                    user_id=current_user.id,
                    diff={"question": question, "category_path": category_path_val},
                )

            existing_questions.add(question)
            success += 1

        except Exception as exc:
            # SAVEPOINT 已自動 rollback 該筆，無需手動 rollback 整批
            errors.append({"row": row_num, "reason": str(exc)})

    db.commit()

    return {
        "success": True,
        "data": {
            "imported": success,
            "skipped": skipped,
            "errors": errors,
            "new_categories": sorted(new_categories),
        },
    }


# ── 匯出端點 ──────────────────────────────────────────────────────────────────

@router.get("/api/v1/agents/{agent_id}/faqs/export")
def export_faqs(
    agent_id: uuid.UUID,
    access: tuple[Agent, str | None] = Depends(get_accessible_agent),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    del access  # 僅做存取驗證

    items = (
        db.query(KnowledgeItem)
        .filter(KnowledgeItem.agent_id == agent_id)
        .order_by(KnowledgeItem.created_at)
        .all()
    )

    # I6：一次性載入該 agent 所有 categories，避免每筆 FAQ 重新遞迴查詢
    all_cats = db.query(Category).filter(Category.agent_id == agent_id).all()
    cat_map: dict[Any, Category] = {c.id: c for c in all_cats}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FAQs"  # type: ignore[union-attr]

    ws.append(  # type: ignore[union-attr]
        ["id", "status", "category_path", "tags", "question", "answer",
         "version", "created_at", "updated_at"]
    )

    for item in items:
        category_path = build_category_path(item.category_id, cat_map)
        tags_str = ",".join(item.tags) if item.tags else ""
        ws.append([  # type: ignore[union-attr]
            str(item.id),
            str(item.status),
            category_path,
            tags_str,
            str(item.question),
            str(item.answer),
            item.version if item.version is not None else 1,
            item.created_at.isoformat() if item.created_at else "",
            item.updated_at.isoformat() if item.updated_at else "",
        ])

    # 稽核紀錄
    record_audit(
        db,
        agent_id=agent_id,
        item_id=None,
        action="export",
        user_id=current_user.id,
        diff={"count": len(items)},
    )
    db.commit()

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    filename = _generate_export_filename("全量", str(agent_id))
    filename_encoded = quote(filename, encoding="utf-8", safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
        },
    )


# ── 分類匯出端點 ──────────────────────────────────────────────────────────────

@router.get("/api/v1/agents/{agent_id}/categories/{category_id}/export")
def export_category_faqs(
    agent_id: uuid.UUID,
    category_id: uuid.UUID,
    access: tuple[Agent, str | None] = Depends(get_accessible_agent),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    del access  # 僅做存取驗證

    category = (
        db.query(Category)
        .filter(Category.id == category_id, Category.agent_id == agent_id)
        .first()
    )
    if category is None:
        raise HTTPException(status_code=404, detail="找不到分類")

    all_cats = db.query(Category).filter(Category.agent_id == agent_id).all()
    cat_map: dict[Any, Category] = {c.id: c for c in all_cats}

    # 計算選定分類的路徑（用於檔名）
    selected_cat_path = build_category_path(category_id, cat_map) or f"category_{category_id}"

    cat_ids = collect_category_subtree(category_id, cat_map)

    items = (
        db.query(KnowledgeItem)
        .filter(KnowledgeItem.agent_id == agent_id)
        .filter(KnowledgeItem.category_id.in_(cat_ids))
        .order_by(KnowledgeItem.created_at)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FAQs"  # type: ignore[union-attr]
    ws.append(["question", "answer", "tags", "category_path", "status", "version"])  # type: ignore[union-attr]

    for item in items:
        tags_str = ",".join(item.tags) if item.tags else ""
        item_cat_path = build_category_path(item.category_id, cat_map)
        ws.append([  # type: ignore[union-attr]
            str(item.question),
            str(item.answer),
            tags_str,
            item_cat_path,
            str(item.status),
            item.version if item.version is not None else 1,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    record_audit(
        db,
        agent_id=agent_id,
        item_id=None,
        action="export_category",
        user_id=current_user.id,
        diff={"category_id": str(category_id), "count": len(items)},
    )
    db.commit()

    filename = _generate_export_filename(selected_cat_path, str(agent_id))
    filename_encoded = quote(filename, encoding="utf-8", safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{filename_encoded}"
        },
    )


# ── 分類匯入端點 ──────────────────────────────────────────────────────────────

@router.post("/api/v1/agents/{agent_id}/categories/{category_id}/import")
def import_category_faqs(
    agent_id: uuid.UUID,
    category_id: uuid.UUID,
    mode: Literal["append", "replace"] = Query("append"),
    file: UploadFile = File(...),
    access: tuple[Agent, str | None] = Depends(get_accessible_agent),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
) -> dict[str, Any]:
    del access  # 僅做存取驗證

    # 驗證分類存在且屬於此 agent
    category = (
        db.query(Category)
        .filter(Category.id == category_id, Category.agent_id == agent_id)
        .first()
    )
    if category is None:
        raise HTTPException(status_code=404, detail="找不到分類")

    # replace 模式：先刪除該分類（含所有子分類）的全部 FAQ
    if mode == "replace":
        all_cats = db.query(Category).filter(Category.agent_id == agent_id).all()
        cat_map_del: dict[Any, Category] = {c.id: c for c in all_cats}
        del_cat_ids = collect_category_subtree(category_id, cat_map_del)
        items_to_del = (
            db.query(KnowledgeItem)
            .filter(
                KnowledgeItem.agent_id == agent_id,
                KnowledgeItem.category_id.in_(del_cat_ids),
            )
            .all()
        )
        deleted_count = len(items_to_del)
        for item in items_to_del:
            db.delete(item)
        db.flush()
        if deleted_count > 0:
            record_audit(
                db,
                agent_id=agent_id,
                item_id=None,
                action="bulk_delete_category",
                user_id=current_user.id,
                diff={"category_id": str(category_id), "deleted_count": deleted_count},
            )

    # 檔案格式與大小驗證
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="僅接受 .xlsx 格式檔案")

    content = file.file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="檔案大小超過 10 MB 上限")

    # 解析 Excel
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("找不到工作表")
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"無法解析 Excel 檔案：{exc}") from exc

    if len(all_rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 無資料列（需含標題行）")

    header = [str(c).strip().lower() if c is not None else "" for c in all_rows[0]]
    data_rows = all_rows[1:]

    if len(data_rows) > MAX_ROWS:
        raise HTTPException(status_code=400, detail=f"資料列數超過 {MAX_ROWS} 行上限")

    # 必填：question、answer；選填：tags、category_path
    def require_col(name: str) -> int:
        if name not in header:
            raise HTTPException(status_code=400, detail=f"缺少必填欄位：{name}")
        return header.index(name)

    q_idx = require_col("question")
    a_idx = require_col("answer")
    tags_idx: Optional[int] = header.index("tags") if "tags" in header else None
    # category_path 選填：有則覆寫 URL 指定分類，自動建立不存在的節點
    cp_idx: Optional[int] = header.index("category_path") if "category_path" in header else None

    # 查詢現有問題以避免重複
    # append 模式：若檔案含 category_path（項目可分散至任意分類），擴大至整個 agent；
    #             否則只檢查 URL 指定分類的子樹範圍
    # replace 模式：子樹已清空，僅追蹤本次操作內的重複
    if mode == "replace":
        existing_questions: set[str] = set()
    elif cp_idx is not None:
        # 檔案含 category_path，各列可路由至不同分類，以全 agent 範圍去重
        existing_questions = {
            str(q)
            for (q,) in db.query(KnowledgeItem.question)
            .filter(KnowledgeItem.agent_id == agent_id)
            .all()
        }
    else:
        # 無 category_path，只檢查 URL 指定分類子樹
        all_cats_append = db.query(Category).filter(Category.agent_id == agent_id).all()
        cat_map_append: dict[Any, Category] = {c.id: c for c in all_cats_append}
        append_cat_ids = collect_category_subtree(category_id, cat_map_append)
        existing_questions = {
            str(q)
            for (q,) in db.query(KnowledgeItem.question)
            .filter(
                KnowledgeItem.agent_id == agent_id,
                KnowledgeItem.category_id.in_(append_cat_ids),
            )
            .all()
        }

    # 預載分類索引（同 import_faqs 邏輯，僅在檔案含 category_path 時用得到）
    category_cache: dict[tuple[Optional[Any], str], Any] = _preload_category_index(
        db, agent_id
    )

    # Superadmin 匯入直接核准（與 create_faq 行為一致），其他使用者建立為 draft
    initial_status = "approved" if current_user.is_superadmin else "draft"

    success = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    for row_num, row in enumerate(data_rows, start=2):
        def cell(idx: int) -> str:
            val = row[idx] if idx < len(row) else None  # type: ignore[index]
            return str(val).strip() if val is not None else ""

        question = cell(q_idx)
        answer = cell(a_idx)
        tags_raw = cell(tags_idx) if tags_idx is not None else ""
        category_path_val = cell(cp_idx) if cp_idx is not None else ""

        # 空列跳過
        if not question and not answer:
            continue

        if not question or not answer:
            errors.append({"row": row_num, "reason": "question / answer 不可為空"})
            continue

        if question in existing_questions:
            skipped += 1
            continue

        tags_list = [t.strip() for t in tags_raw.split(",") if t.strip()] if tags_raw else []

        try:
            with db.begin_nested():
                # category_path 優先：有值則解析路徑（自動建立缺失節點），否則使用 URL 指定分類
                if category_path_val:
                    target_cat_id, _ = _resolve_category_path(
                        db, agent_id, category_path_val, cache=category_cache
                    )
                else:
                    target_cat_id = category_id

                item = KnowledgeItem(
                    id=uuid.uuid4(),
                    agent_id=agent_id,
                    category_id=target_cat_id,
                    question=question,
                    answer=answer,
                    tags=tags_list,
                    status=initial_status,
                    version=1,
                    created_by=current_user.id,
                )
                db.add(item)
                db.flush()

                record_audit(
                    db,
                    agent_id=agent_id,
                    item_id=item.id,
                    action="import_category",
                    user_id=current_user.id,
                    diff={
                        "question": question,
                        "category_id": str(target_cat_id),
                        "category_path": category_path_val,
                    },
                )

            existing_questions.add(question)
            success += 1

        except Exception as exc:
            errors.append({"row": row_num, "reason": str(exc)})

    db.commit()

    return {
        "success": True,
        "data": {
            "imported": success,
            "skipped": skipped,
            "errors": errors,
        },
    }
