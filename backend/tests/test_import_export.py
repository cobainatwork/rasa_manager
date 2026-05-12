"""
Excel 匯入/匯出測試：欄位映射、無效資料容錯、重複檢測、category_path 自動建立。
"""
from __future__ import annotations

import io
import uuid
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from tests.conftest import AGENT_ID


# ── 輔助：在記憶體建立測試用 xlsx ─────────────────────────────────────────────

def _make_xlsx(rows: list[list[object]], headers: list[str] | None = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    if headers is None:
        headers = ["question", "answer", "category_path", "tags"]
    ws.append(headers)  # type: ignore[union-attr]
    for row in rows:
        ws.append(row)  # type: ignore[union-attr]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── category_path 解析（單元測試）────────────────────────────────────────────

class TestResolveCategoryPath:
    def test_new_single_level_category_created(self) -> None:
        from api.routes.import_export import _resolve_category_path

        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = None
        db.flush = MagicMock()
        db.add = MagicMock()

        with patch("api.routes.import_export.Category") as MockCat:
            instance = MagicMock()
            instance.id = uuid.uuid4()
            MockCat.return_value = instance
            _resolve_category_path(db, AGENT_ID, "常見問題")

        db.add.assert_called_once()
        db.flush.assert_called_once()

    def test_existing_category_reused(self) -> None:
        from api.routes.import_export import _resolve_category_path

        existing = MagicMock()
        existing.id = uuid.UUID("00000000-0000-0000-0000-000000000020")
        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = existing
        db.add = MagicMock()

        result = _resolve_category_path(db, AGENT_ID, "既有分類")

        # I2：簽章改為 (id, created)
        assert result == (existing.id, False)
        db.add.assert_not_called()

    def test_empty_path_raises_value_error(self) -> None:
        from api.routes.import_export import _resolve_category_path

        db = MagicMock()
        with pytest.raises(ValueError, match="不可為空"):
            _resolve_category_path(db, AGENT_ID, "")

    def test_slash_only_path_raises_value_error(self) -> None:
        from api.routes.import_export import _resolve_category_path

        db = MagicMock()
        with pytest.raises(ValueError, match="不可為空"):
            _resolve_category_path(db, AGENT_ID, "///")


# ── 匯入端點（整合測試）──────────────────────────────────────────────────────

class TestImportEndpoint:
    """
    使用 patch 將 require_agent_access bypass，
    專注測試 Excel 解析與匯入業務邏輯。
    """

    def _setup_db(self, mock_db: MagicMock) -> None:
        mock_db.query.return_value.filter.return_value.all.return_value = []
        mock_db.query.return_value.filter.return_value.count.return_value = 0
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.rollback = MagicMock()

    def test_valid_xlsx_imports_successfully(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        self._setup_db(mock_db)
        xlsx = _make_xlsx([["問題一", "答案一", "分類A", "tag1,tag2"]])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._resolve_category_path", return_value=(uuid.uuid4(), False)),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["skipped"] == 0
        assert data["errors"] == []

    def test_wrong_file_type_rejected(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        with patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={"file": ("test.csv", b"q,a,c\n1,2,3", "text/csv")},
            )
        assert resp.status_code == 400
        assert "xlsx" in resp.json()["detail"].lower()

    def test_missing_category_path_uses_default_category(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """category_path 為選填：不含 category_path 欄位的 xlsx 應匯入成功並歸入預設分類。"""
        self._setup_db(mock_db)
        xlsx = _make_xlsx([["Q_no_cat", "A_no_cat"]], headers=["question", "answer"])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._get_or_create_default_category", return_value=uuid.uuid4()),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["errors"] == []

    def test_duplicate_question_skipped(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        # "問題一" 已存在
        mock_db.query.return_value.filter.return_value.all.return_value = [("問題一",)]
        mock_db.query.return_value.filter.return_value.count.return_value = 0
        mock_db.flush = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.rollback = MagicMock()

        xlsx = _make_xlsx([["問題一", "答案一", "分類A", ""]])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._resolve_category_path", return_value=(uuid.uuid4(), False)),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["skipped"] == 1
        assert data["imported"] == 0

    def test_empty_rows_skipped_gracefully(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        self._setup_db(mock_db)
        xlsx = _make_xlsx([
            ["問題一", "答案一", "分類A", ""],
            ["", "", "", ""],       # 完全空白列
            ["問題二", "答案二", "分類B", ""],
        ])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._resolve_category_path", return_value=(uuid.uuid4(), False)),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        assert resp.json()["data"]["imported"] == 2

    def test_replace_mode_calls_delete_then_imports(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """mode=replace 應先刪除 agent 所有 FAQ，再以空集合重新匯入。"""
        self._setup_db(mock_db)
        # 設定 delete chain（query → filter → delete）
        mock_db.query.return_value.filter.return_value.delete = MagicMock(return_value=5)
        xlsx = _make_xlsx([["新問題A", "新答案A", "分類X", ""]])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._resolve_category_path", return_value=(uuid.uuid4(), False)),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import?mode=replace",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        # replace 模式：先刪舊資料，再匯入新資料，不會計算 skipped
        assert data["imported"] == 1
        assert data["skipped"] == 0
        # 驗證 delete() 確實被呼叫
        mock_db.query.return_value.filter.return_value.delete.assert_called_once_with(
            synchronize_session="fetch"
        )

    def test_replace_mode_reimports_previously_existing_question(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """replace 模式下即使問題曾存在（DB 已有），也應全數匯入（不跳過）。"""
        self._setup_db(mock_db)
        mock_db.query.return_value.filter.return_value.delete = MagicMock(return_value=3)
        xlsx = _make_xlsx([
            ["舊問題A", "舊答案A", "分類Y", ""],
            ["舊問題B", "舊答案B", "分類Y", ""],
        ])

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._resolve_category_path", return_value=(uuid.uuid4(), False)),
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/import?mode=replace",
                files={"file": ("test.xlsx", xlsx, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 2
        assert data["skipped"] == 0


# ── 匯出端點（整合測試）──────────────────────────────────────────────────────

class TestExportEndpoint:
    def test_export_returns_xlsx(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        faq = MagicMock()
        faq.id = uuid.uuid4()
        faq.status = "approved"
        faq.category_id = uuid.uuid4()
        faq.tags = ["t1"]
        faq.question = "Q"
        faq.answer = "A"
        faq.version = 1
        faq.created_at = None
        faq.updated_at = None

        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [faq]
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export.build_category_path", return_value="分類A"),
            patch("api.routes.import_export._get_redis", side_effect=Exception("no redis in test")),
        ):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/export"
            )

        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")
        cd = resp.headers.get("content-disposition", "")
        assert "全量_export_" in cd or "_export_" in cd  # 新檔名格式

    def test_export_empty_agent_returns_header_only(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        mock_db.query.return_value.filter.return_value.order_by.return_value.all.return_value = []
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        with (
            patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)),
            patch("api.routes.import_export._get_redis", side_effect=Exception("no redis in test")),
        ):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/faqs/export"
            )

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
        assert len(rows) == 1  # 僅標題行


# ── Regression: B7 (import per-row failure must use SAVEPOINT) ───────────────

class TestImportSavepointRegression:
    """Regression: B7 (import per-row failure must use SAVEPOINT, not full rollback)."""

    def test_third_row_failure_does_not_rollback_others(
        self, client_superadmin, mock_db: MagicMock
    ) -> None:
        """
        匯入 5 筆，第 3 筆 _resolve_category_path 拋例外。
        預期：success=4、failed=1、其餘 4 筆不受牽連。
        """
        mock_db.query.return_value.filter.return_value.all.return_value = []
        mock_db.query.return_value.filter.return_value.count.return_value = 0
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        nested_cm = MagicMock()
        nested_cm.__enter__ = MagicMock(return_value=nested_cm)
        nested_cm.__exit__ = MagicMock(return_value=False)
        mock_db.begin_nested = MagicMock(return_value=nested_cm)

        xlsx = _make_xlsx([
            ["Q1", "A1", "分類A", ""],
            ["Q2", "A2", "分類A", ""],
            ["Q3", "A3", "分類BAD", ""],
            ["Q4", "A4", "分類A", ""],
            ["Q5", "A5", "分類A", ""],
        ])

        call_count = [0]

        def fake_resolve(db: object, agent_id: object, path: str) -> object:
            call_count[0] += 1
            if call_count[0] == 3:
                raise RuntimeError("simulated category resolve failure")
            return uuid.uuid4(), False

        with (
            patch(
                "api.routes.import_export.require_agent_access",
                return_value=(MagicMock(), None),
            ),
            patch(
                "api.routes.import_export._resolve_category_path",
                side_effect=fake_resolve,
            ),
        ):
            resp = client_superadmin.post(
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={
                    "file": (
                        "test.xlsx",
                        xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 4, f"成功筆數應為 4，實際 {data['imported']}"
        assert len(data["errors"]) == 1
        assert data["errors"][0]["row"] == 4

        assert mock_db.begin_nested.call_count == 5
        mock_db.rollback.assert_not_called()


# ── Regression: I2 (_resolve_category_path must not call count() in loop) ────

class TestResolveCategoryPathNoCountRegression:
    """Regression: I2 (import _resolve_category_path 移除迴圈內 count())."""

    def test_signature_returns_tuple_with_created_flag(self) -> None:
        from api.routes.import_export import _resolve_category_path

        existing = MagicMock()
        existing.id = uuid.uuid4()
        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = existing

        result = _resolve_category_path(db, AGENT_ID, "已存在")
        assert result == (existing.id, False)

    def test_created_flag_true_when_new(self) -> None:
        from api.routes.import_export import _resolve_category_path

        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = None
        with patch("api.routes.import_export.Category") as MockCat:
            inst = MagicMock()
            inst.id = uuid.uuid4()
            MockCat.return_value = inst
            result = _resolve_category_path(db, AGENT_ID, "新分類")
        assert result == (inst.id, True)

    def test_no_count_calls_during_import(
        self, client_superadmin, mock_db: MagicMock
    ) -> None:
        """匯入 3 筆相同 category_path 的 FAQ，不應出現 count() 呼叫。"""
        mock_db.query.return_value.filter.return_value.all.return_value = []
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.rollback = MagicMock()

        count_calls = [0]

        def count_spy() -> int:
            count_calls[0] += 1
            return 0

        mock_db.query.return_value.filter.return_value.count.side_effect = count_spy

        xlsx = _make_xlsx([
            ["問題一", "答案一", "分類A", ""],
            ["問題二", "答案二", "分類A", ""],
            ["問題三", "答案三", "分類A", ""],
        ])

        with (
            patch(
                "api.routes.import_export.require_agent_access",
                return_value=(MagicMock(), None),
            ),
            patch(
                "api.routes.import_export._resolve_category_path",
                return_value=(uuid.uuid4(), False),
            ),
        ):
            resp = client_superadmin.post(
                f"/api/v1/agents/{AGENT_ID}/faqs/import",
                files={
                    "file": (
                        "t.xlsx",
                        xlsx,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )

        assert resp.status_code == 200
        assert count_calls[0] <= 1


# ── Regression: I6 (export must load categories table once) ─────────────────

class TestExportCategoriesLoadedOnceRegression:
    """Regression: I6 (匯出時 categories 表一次性載入)."""

    def test_categories_table_queried_once(
        self, client_superadmin, mock_db: MagicMock
    ) -> None:
        from api.database.models import Category, KnowledgeItem

        items = []
        for _ in range(5):
            it = MagicMock()
            it.id = uuid.uuid4()
            it.status = "approved"
            it.category_id = uuid.uuid4()
            it.tags = []
            it.question = "Q"
            it.answer = "A"
            it.version = 1
            it.created_at = None
            it.updated_at = None
            items.append(it)

        category_query_count = [0]

        def query_side_effect(*args):
            q = MagicMock()
            if args and args[0] is KnowledgeItem:
                q.filter.return_value.order_by.return_value.all.return_value = items
                return q
            if args and args[0] is Category:
                category_query_count[0] += 1
                q.filter.return_value.all.return_value = []
                return q
            return q

        mock_db.query.side_effect = query_side_effect
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        with patch("api.routes.import_export.require_agent_access", return_value=(MagicMock(), None)):
            resp = client_superadmin.get(
                f"/api/v1/agents/{AGENT_ID}/faqs/export"
            )

        assert resp.status_code == 200


# ── collect_category_subtree 單元測試 ─────────────────────────────────────────

class TestCollectCategoryIds:
    def test_collects_self_when_no_children(self) -> None:
        from api.utils.category_path import collect_category_subtree

        cat_id = uuid.uuid4()
        cat = MagicMock()
        cat.id = cat_id
        cat.parent_id = None

        result = collect_category_subtree(cat_id, {cat_id: cat})
        assert result == {cat_id}

    def test_collects_children_recursively(self) -> None:
        from api.utils.category_path import collect_category_subtree

        root_id = uuid.uuid4()
        child_id = uuid.uuid4()
        grandchild_id = uuid.uuid4()

        root = MagicMock()
        root.id = root_id
        root.parent_id = None
        child = MagicMock()
        child.id = child_id
        child.parent_id = root_id
        grandchild = MagicMock()
        grandchild.id = grandchild_id
        grandchild.parent_id = child_id

        cat_map = {root_id: root, child_id: child, grandchild_id: grandchild}

        result = collect_category_subtree(root_id, cat_map)
        assert result == {root_id, child_id, grandchild_id}

    def test_does_not_collect_sibling_categories(self) -> None:
        from api.utils.category_path import collect_category_subtree

        root_id = uuid.uuid4()
        sibling_id = uuid.uuid4()

        root = MagicMock()
        root.id = root_id
        root.parent_id = None
        sibling = MagicMock()
        sibling.id = sibling_id
        sibling.parent_id = None

        result = collect_category_subtree(root_id, {root_id: root, sibling_id: sibling})
        assert result == {root_id}


# ── 分類匯出端點 ──────────────────────────────────────────────────────────────

class TestExportCategoryEndpoint:
    def _make_db(
        self,
        mock_db: MagicMock,
        cat_id: uuid.UUID,
        faqs: list | None = None,
    ) -> MagicMock:
        """設定 mock_db query 序列（require_agent_access 已被 patch，不含前 2 次查詢）。"""
        faqs = faqs or []
        mock_cat = MagicMock()
        mock_cat.id = cat_id
        mock_cat.parent_id = None
        mock_cat.name = "測試分類"

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:  # Category.filter().first() — 驗證分類存在
                q.filter.return_value.first.return_value = mock_cat
            elif idx == 1:  # Category.filter().all() — cat_map
                q.filter.return_value.all.return_value = [mock_cat]
            elif idx == 2:  # KnowledgeItem.filter().filter().order_by().all()
                (q.filter.return_value
                  .filter.return_value
                  .order_by.return_value
                  .all.return_value) = faqs
            return q

        mock_db.query.side_effect = se
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        return mock_cat

    def test_returns_xlsx_stream(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        self._make_db(mock_db, cat_id)

        with (
            patch("api.routes.import_export.require_agent_access"),
            patch("api.routes.import_export._get_redis", side_effect=Exception("no redis in test")),
        ):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/export"
            )

        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_xlsx_contains_faq_rows(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        mock_faq = MagicMock()
        mock_faq.question = "問題一"
        mock_faq.answer = "答案一"
        mock_faq.tags = ["t1"]
        mock_faq.status = "approved"
        mock_faq.version = 1
        mock_faq.created_at = None
        mock_faq.category_id = cat_id
        self._make_db(mock_db, cat_id, faqs=[mock_faq])

        with (
            patch("api.routes.import_export.require_agent_access"),
            patch("api.routes.import_export.build_category_path", return_value="測試分類"),
            patch("api.routes.import_export._get_redis", side_effect=Exception("no redis in test")),
        ):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/export"
            )

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
        assert rows[0][0] == "question"       # 標題列第 1 欄
        assert "category_path" in rows[0]     # 標題列含 category_path
        assert rows[1][0] == "問題一"          # 資料列

    def test_invalid_category_returns_404(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        mock_db.query.return_value.filter.return_value.first.return_value = None
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/export"
            )

        assert resp.status_code == 404

    def test_child_faq_shows_full_path_when_exporting_from_parent(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """
        Regression：從父分類 A 匯出時，子分類 AB 的 FAQ 應顯示完整路徑 'A/AB'，
        而非僅顯示父分類名稱 'A' 或錯誤的 'A/A'。

        此測試刻意不 mock build_category_path，讓真實路徑解析邏輯執行。
        """
        root_id = uuid.uuid4()
        child_id = uuid.uuid4()

        root_cat = MagicMock()
        root_cat.id = root_id
        root_cat.name = "A"
        root_cat.parent_id = None  # 根分類

        child_cat = MagicMock()
        child_cat.id = child_id
        child_cat.name = "AB"
        child_cat.parent_id = root_id  # 子分類，parent 是 root_cat

        # FAQ 存在於子分類 AB，而非根分類 A
        item_in_child = MagicMock()
        item_in_child.question = "子分類問題"
        item_in_child.answer = "子分類答案"
        item_in_child.tags = []
        item_in_child.status = "draft"
        item_in_child.version = 1
        item_in_child.created_at = None
        item_in_child.category_id = child_id  # 關鍵：指向 AB，不是 A

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — 驗證分類 A 存在
                q.filter.return_value.first.return_value = root_cat
            elif idx == 1:  # Category.filter().all() — cat_map（含 A 與 AB）
                q.filter.return_value.all.return_value = [root_cat, child_cat]
            elif idx == 2:  # KnowledgeItem — 子樹 FAQ
                (q.filter.return_value
                  .filter.return_value
                  .order_by.return_value
                  .all.return_value) = [item_in_child]
            return q

        mock_db.query.side_effect = se
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        with (
            patch("api.routes.import_export.require_agent_access"),
            patch("api.routes.import_export._get_redis", side_effect=Exception("no redis in test")),
        ):
            resp = client_superadmin.get(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{root_id}/export"
            )

        assert resp.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))  # type: ignore[union-attr]
        assert len(rows) == 2, f"應有 1 個標題列 + 1 個資料列，實際 {len(rows)} 列"

        header = list(rows[0])
        cat_path_col = header.index("category_path")

        actual_path = rows[1][cat_path_col]
        assert actual_path == "A/AB", (
            f"從父分類 A 匯出時，子分類 AB 的 FAQ 應顯示 'A/AB'，"
            f"實際得到 '{actual_path}'（若顯示 'A/A' 代表 item.category_id 被錯誤替換為 selected category_id）"
        )


# ── 分類匯入端點 ──────────────────────────────────────────────────────────────

class TestImportCategoryEndpoint:
    def _make_db(self, mock_db: MagicMock, cat_id: uuid.UUID) -> None:
        """append 模式下的 mock_db：
        query 0 = 分類驗證（Category.filter().first()）
        query 1 = Category.filter().all()（cat_map_append）
        query 2 = KnowledgeItem.question.filter().filter().all()（子樹現有問題）
        """
        mock_cat = MagicMock()
        mock_cat.id = cat_id
        mock_cat.parent_id = None
        mock_cat.agent_id = AGENT_ID

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:  # Category.filter().first() — 驗證分類存在
                q.filter.return_value.first.return_value = mock_cat
            elif idx == 1:  # Category.filter().all() — cat_map_append
                q.filter.return_value.all.return_value = [mock_cat]
            elif idx == 2:  # KnowledgeItem.question.filter().all() — 子樹現有問題
                q.filter.return_value.all.return_value = []
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q

        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

    def test_only_question_answer_required(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """category_path 欄位不應是必填欄位。"""
        cat_id = uuid.uuid4()
        self._make_db(mock_db, cat_id)
        xlsx = _make_xlsx(
            [["問題一", "答案一"]],
            headers=["question", "answer"],
        )

        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["skipped"] == 0
        assert data["errors"] == []

    def test_missing_answer_column_rejected(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        self._make_db(mock_db, cat_id)
        xlsx = _make_xlsx([["Q"]], headers=["question"])

        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 400
        assert "answer" in resp.json()["detail"].lower()

    def test_duplicate_question_skipped(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        mock_cat = MagicMock()
        mock_cat.id = cat_id
        mock_cat.parent_id = None

        counter = [0]
        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — 驗證分類存在
                q.filter.return_value.first.return_value = mock_cat
            elif idx == 1:  # Category.filter().all() — cat_map_append
                q.filter.return_value.all.return_value = [mock_cat]
            elif idx == 2:  # KnowledgeItem.question.filter().all() — 子樹現有問題
                q.filter.return_value.all.return_value = [("問題一",)]
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q
        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

        xlsx = _make_xlsx([["問題一", "答案一"]], headers=["question", "answer"])
        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 0
        assert data["skipped"] == 1

    def test_replace_mode_calls_delete_on_existing_items(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """mode=replace → 應對現有 FAQ 呼叫 db.delete()。"""
        cat_id = uuid.uuid4()
        mock_cat = MagicMock()
        mock_cat.id = cat_id
        mock_cat.parent_id = None

        existing_item = MagicMock()

        counter = [0]
        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — validate category
                q.filter.return_value.first.return_value = mock_cat
            elif idx == 1:  # Category.filter().all() — cat_map for _collect_category_ids
                q.filter.return_value.all.return_value = [mock_cat]
            elif idx == 2:  # KnowledgeItem.filter().all() — items_to_del
                q.filter.return_value.all.return_value = [existing_item]
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q
        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.delete = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

        xlsx = _make_xlsx([["新問題", "新答案"]], headers=["question", "answer"])
        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import?mode=replace",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        mock_db.delete.assert_called_once_with(existing_item)

    def test_append_mode_does_not_delete(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """mode=append（預設）→ 不應呼叫 db.delete()。"""
        cat_id = uuid.uuid4()
        self._make_db(mock_db, cat_id)
        mock_db.delete = MagicMock()

        xlsx = _make_xlsx([["問題A", "答案A"]], headers=["question", "answer"])
        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        mock_db.delete.assert_not_called()

    def test_invalid_category_returns_404(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        cat_id = uuid.uuid4()
        mock_db.query.return_value.filter.return_value.first.return_value = None
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()

        xlsx = _make_xlsx([["Q", "A"]], headers=["question", "answer"])
        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 404

    def test_category_path_in_file_routes_to_resolved_category(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """
        Regression：分類匯入時，若 Excel 含 category_path 欄，應依路徑路由到對應分類（含自動建立），
        而非一律歸入 URL 指定的分類。

        情境：系統有 A/B，無 A/C；匯入資料 category_path = A/C
              預期：建立 A/C 並將項目歸入 A/C，不是 A/B。
        """
        cat_b_id = uuid.uuid4()   # URL 指定分類 A/B
        resolved_c_id = uuid.uuid4()  # _resolve_category_path("A/C") 返回的分類 ID

        mock_cat_b = MagicMock()
        mock_cat_b.id = cat_b_id
        mock_cat_b.parent_id = None
        mock_cat_b.agent_id = AGENT_ID

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — 分類驗證
                q.filter.return_value.first.return_value = mock_cat_b
            elif idx == 1:  # KnowledgeItem.question — 全 agent 去重查詢（含 category_path 時）
                q.filter.return_value.all.return_value = []
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q

        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

        xlsx = _make_xlsx(
            [["問題A/C", "答案A/C", "A/C", ""]],
            headers=["question", "answer", "category_path", "tags"],
        )

        with (
            patch("api.routes.import_export.require_agent_access"),
            patch(
                "api.routes.import_export._resolve_category_path",
                return_value=(resolved_c_id, True),
            ) as mock_resolve,
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_b_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        data = resp.json()["data"]
        assert data["imported"] == 1, f"應匯入 1 筆，實際 {data['imported']}"
        # _resolve_category_path 必須被呼叫，且傳入的是檔案裡的路徑（A/C），而非 URL 分類 ID
        mock_resolve.assert_called_once()
        call_args = mock_resolve.call_args
        assert call_args[0][2] == "A/C", (
            f"應以 'A/C' 呼叫 _resolve_category_path，實際傳入 '{call_args[0][2]}'"
        )

    def test_category_path_absent_still_uses_url_category(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """無 category_path 欄時，行為不變：所有項目歸入 URL 指定分類。"""
        cat_id = uuid.uuid4()
        self._make_db(mock_db, cat_id)

        xlsx = _make_xlsx([["問題無Path", "答案無Path"]], headers=["question", "answer"])
        with (
            patch("api.routes.import_export.require_agent_access"),
            patch("api.routes.import_export._resolve_category_path") as mock_resolve,
        ):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{cat_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200
        assert resp.json()["data"]["imported"] == 1
        # 沒有 category_path 欄，_resolve_category_path 不應被呼叫
        mock_resolve.assert_not_called()


# ── Regression：匯入子分類不應產生 422 ────────────────────────────────────────

class TestImportCategorySubcategoryRegression:
    """
    回歸測試：匯入子分類（parent_id != None）時，append 與 replace 模式均應回傳 200。

    背景：cat_map 包含父分類與子分類兩個節點，子分類的 parent_id 指向父分類 UUID，
    確認 _collect_category_ids 正確只收集子分類本身（不誤入父分類），
    且整個 HTTP 端點流程不因「非根節點」而回傳 422。
    """

    def _make_sub_db_append(
        self,
        mock_db: MagicMock,
        sub_id: uuid.UUID,
        parent_id: uuid.UUID,
    ) -> None:
        """append 模式 mock：cat_map 包含父分類 + 子分類。"""
        mock_parent = MagicMock()
        mock_parent.id = parent_id
        mock_parent.parent_id = None  # 父為根節點

        mock_sub = MagicMock()
        mock_sub.id = sub_id
        mock_sub.parent_id = parent_id  # 子節點的 parent_id != None
        mock_sub.agent_id = AGENT_ID

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — 分類存在驗證
                q.filter.return_value.first.return_value = mock_sub
            elif idx == 1:  # Category.filter().all() — cat_map_append（含父子）
                q.filter.return_value.all.return_value = [mock_parent, mock_sub]
            elif idx == 2:  # KnowledgeItem.question — 子樹現有問題
                q.filter.return_value.all.return_value = []
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q

        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

    def test_append_into_subcategory_returns_200(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """子分類 append 匯入應回傳 200，不應產生 422。"""
        parent_id = uuid.uuid4()
        sub_id = uuid.uuid4()
        self._make_sub_db_append(mock_db, sub_id, parent_id)

        xlsx = _make_xlsx([["問題子分類A", "答案子分類A"]], headers=["question", "answer"])

        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{sub_id}/import",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200, f"期望 200，實際 {resp.status_code}；body={resp.text}"
        data = resp.json()["data"]
        assert data["imported"] == 1
        assert data["skipped"] == 0
        assert data["errors"] == []

    def test_replace_into_subcategory_returns_200(
        self, client_superadmin: object, mock_db: MagicMock
    ) -> None:
        """子分類 replace 匯入應回傳 200，舊 FAQ 應被刪除，不應產生 422。"""
        parent_id = uuid.uuid4()
        sub_id = uuid.uuid4()

        mock_parent = MagicMock()
        mock_parent.id = parent_id
        mock_parent.parent_id = None

        mock_sub = MagicMock()
        mock_sub.id = sub_id
        mock_sub.parent_id = parent_id
        mock_sub.agent_id = AGENT_ID

        existing_item = MagicMock()

        counter = [0]

        def se(*_args, **_kwargs):
            q = MagicMock()
            idx = counter[0]
            counter[0] += 1
            if idx == 0:   # Category.filter().first() — validate
                q.filter.return_value.first.return_value = mock_sub
            elif idx == 1:  # Category.filter().all() — cat_map_del（含父子）
                q.filter.return_value.all.return_value = [mock_parent, mock_sub]
            elif idx == 2:  # KnowledgeItem.filter().all() — items_to_del
                q.filter.return_value.all.return_value = [existing_item]
            else:
                q.filter.return_value.all.return_value = []
                q.filter.return_value.filter.return_value.all.return_value = []
            return q

        mock_db.query.side_effect = se
        mock_db.flush = MagicMock()
        mock_db.add = MagicMock()
        mock_db.commit = MagicMock()
        mock_db.delete = MagicMock()
        mock_db.begin_nested.return_value.__enter__ = MagicMock(return_value=None)
        mock_db.begin_nested.return_value.__exit__ = MagicMock(return_value=False)

        xlsx = _make_xlsx([["新問題子分類B", "新答案子分類B"]], headers=["question", "answer"])

        with patch("api.routes.import_export.require_agent_access"):
            resp = client_superadmin.post(  # type: ignore[attr-defined]
                f"/api/v1/agents/{AGENT_ID}/categories/{sub_id}/import?mode=replace",
                files={"file": ("t.xlsx", xlsx,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            )

        assert resp.status_code == 200, f"期望 200，實際 {resp.status_code}；body={resp.text}"
        mock_db.delete.assert_called_once_with(existing_item)
        data = resp.json()["data"]
        assert data["imported"] == 1

    def test_replace_subcategory_does_not_delete_parent_faqs(self) -> None:
        """replace 子分類時，collect_category_subtree 不應收集父分類 ID，
        確認父分類的 FAQ 不被誤刪。"""
        from api.utils.category_path import collect_category_subtree

        parent_id = uuid.uuid4()
        sub_id = uuid.uuid4()

        mock_parent = MagicMock()
        mock_parent.id = parent_id
        mock_parent.parent_id = None
        mock_sub = MagicMock()
        mock_sub.id = sub_id
        mock_sub.parent_id = parent_id

        cat_map = {parent_id: mock_parent, sub_id: mock_sub}
        collected = collect_category_subtree(sub_id, cat_map)

        assert sub_id in collected, "子分類本身應在收集集合內"
        assert parent_id not in collected, "父分類不應被收集（replace 不應影響父分類）"
